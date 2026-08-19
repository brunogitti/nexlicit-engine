# Geração determinística da planilha de preço (Fase 4, Camada 1, decisão B
# -- 16/08/2026) -- monta um XLSX a partir do catálogo mínimo de itens
# (numero + texto_bruto, extraído sem IA por app/extracao/tabela_itens.py)
# cruzado com o preço digitado por gente (app.db.repositorio.preco_item).
# Sem chamada de IA nenhuma, mesmo princípio de app/geracao/declaracoes.py:
# montagem pura de documento em cima de dado já confiável.

from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

# Mesmo --stamp do CSS do app, na intensidade "levemente colorido" já
# usada em .resposta-nao-encontrada (rgba(139,46,46,0.12) sobre fundo
# branco) -- calculado uma vez, não uma cor nova inventada só pra isto.
# ARGB: "FF" de opacidade total + os 6 dígitos da cor.
_COR_FALTANTE = PatternFill(start_color="FFF1E6E6", end_color="FFF1E6E6", fill_type="solid")

_CABECALHO = ["Item", "Descrição (texto bruto do edital)", "Quantidade", "Preço unitário (R$)", "Preço total (R$)"]


class SemCatalogoError(Exception):
    """O processo não tem catálogo de itens salvo -- ou nunca foi
    (re)processado depois da funcionalidade existir, ou o edital não tem
    tabela de itens reconhecível (ver app.extracao.tabela_itens)."""


def _formatar_cabecalho(planilha: Worksheet) -> None:
    for celula in planilha[1]:
        celula.font = Font(bold=True)


def gerar_planilha_preco(
    processo: dict[str, Any],
    catalogo: list[dict[str, Any]],
    precos: dict[int, dict[str, Any]],
) -> Workbook:
    """Monta o XLSX de planilha de preço pro `processo`, um item por
    linha, na ordem do catálogo. `precos` é um dict numero_item -> linha
    de preco_item (formato de app.db.repositorio.obter_precos_item()).

    Levanta SemCatalogoError se `catalogo` estiver vazio.

    Item sem quantidade OU sem preço unitário: célula(s) correspondentes
    ficam vazias (None) e com destaque visual (fundo rosado, mesmo --stamp
    do resto do app) -- nunca um valor inventado, nunca trava a geração
    dos outros itens. Preço total só é calculado quando os dois números
    existem; sem eles, fica vazio também (não dá pra inventar metade de
    uma conta).
    """
    if not catalogo:
        raise SemCatalogoError(
            f"processo {processo['id']} não tem catálogo de itens salvo -- rode a análise "
            "(ou reprocesse) depois de confirmar que o edital tem uma tabela de itens "
            "reconhecível, ou preencha a planilha manualmente"
        )

    pasta = Workbook()
    planilha = pasta.active
    assert planilha is not None
    planilha.title = "Planilha de Preço"

    planilha.append(_CABECALHO)
    _formatar_cabecalho(planilha)

    total_geral = 0.0
    tem_total_geral = False

    for item in catalogo:
        preco = precos.get(item["numero"])
        quantidade = preco["quantidade"] if preco else None
        preco_unitario = preco["preco_unitario"] if preco else None

        if quantidade is not None and preco_unitario is not None:
            preco_total = quantidade * preco_unitario
            total_geral += preco_total
            tem_total_geral = True
        else:
            preco_total = None

        planilha.append([item["numero"], item["texto_bruto"], quantidade, preco_unitario, preco_total])
        linha_atual = planilha.max_row

        if quantidade is None:
            planilha.cell(row=linha_atual, column=3).fill = _COR_FALTANTE
        if preco_unitario is None:
            planilha.cell(row=linha_atual, column=4).fill = _COR_FALTANTE
        if preco_total is None:
            planilha.cell(row=linha_atual, column=5).fill = _COR_FALTANTE

    planilha.append(["", "", "", "Total geral:", total_geral if tem_total_geral else None])
    linha_total = planilha.max_row
    for celula in planilha[linha_total]:
        celula.font = Font(bold=True)
    if not tem_total_geral:
        planilha.cell(row=linha_total, column=5).fill = _COR_FALTANTE

    planilha.column_dimensions["A"].width = 8
    planilha.column_dimensions["B"].width = 70
    planilha.column_dimensions["C"].width = 14
    planilha.column_dimensions["D"].width = 18
    planilha.column_dimensions["E"].width = 18

    return pasta
