# Testes da planilha de preço (Fase 4, Camada 1, decisão B -- 16/08/2026).
#
# Três frentes: (1) CRUD/upsert do repositório (catálogo + preço), (2)
# geração do XLSX (função pura, sem banco), (3) teste de ouro -- extração
# do catálogo contra um PDF REAL já usado nesta sessão (Paulínia ou
# Frutal, em uploads/), sem chamada de IA nenhuma (fatiar_por_item é
# determinístico), então roda offline, sem gastar cota do Gemini.

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.db.repositorio import (
    RegistroNaoEncontradoError,
    atualizar_validade_proposta,
    criar_processo,
    obter_catalogo_itens,
    obter_precos_item,
    obter_processo,
    salvar_catalogo_itens,
    salvar_preco_item,
)
from app.extracao.extrator import extrair_texto
from app.extracao.tabela_itens import fatiar_por_item, localizar_documento_com_tabela
from app.geracao.planilha_preco import SemCatalogoError, gerar_planilha_preco

PASTA_UPLOADS_REAIS = Path(__file__).resolve().parent.parent / "uploads"
CAMINHO_PAULINIA = PASTA_UPLOADS_REAIS / "7" / "PE 82.2026 - PM PAULINIA - 15.06 - MAXXIMED - BNC - 18974758 - SQ - SIM OK IMPRESSO.pdf"


@pytest.fixture
def caminho_db(tmp_path) -> str:
    return str(tmp_path / "teste.db")


# ---------- Repositório: catálogo (item_catalogo) ----------


def test_salvar_e_obter_catalogo_itens(caminho_db):
    processo_id = criar_processo({"nome": "Processo de teste"}, caminho_banco=caminho_db)

    salvar_catalogo_itens(
        processo_id,
        [
            {"numero": 1, "texto_bruto": "1 Cadeira de rodas UND 5", "pagina": 3, "localizador": "página 3"},
            {"numero": 2, "texto_bruto": "2 Muleta axilar UND 20", "pagina": 3, "localizador": "página 3"},
        ],
        caminho_banco=caminho_db,
    )

    catalogo = obter_catalogo_itens(processo_id, caminho_banco=caminho_db)
    assert [item["numero"] for item in catalogo] == [1, 2]
    assert catalogo[0]["texto_bruto"] == "1 Cadeira de rodas UND 5"
    assert catalogo[0]["pagina"] == 3


def test_obter_catalogo_itens_processo_sem_catalogo_devolve_lista_vazia(caminho_db):
    processo_id = criar_processo({"nome": "Processo sem catálogo"}, caminho_banco=caminho_db)
    assert obter_catalogo_itens(processo_id, caminho_banco=caminho_db) == []


# ---------- Repositório: preço (preco_item, upsert) ----------


def test_salvar_preco_item_cria_na_primeira_vez(caminho_db):
    processo_id = criar_processo({"nome": "Processo de teste"}, caminho_banco=caminho_db)

    linha = salvar_preco_item(processo_id, 1, quantidade=10, preco_unitario=25.5, caminho_banco=caminho_db)

    assert linha["numero_item"] == 1
    assert linha["quantidade"] == 10
    assert linha["preco_unitario"] == 25.5


def test_salvar_preco_item_atualiza_na_segunda_vez_sem_duplicar(caminho_db):
    processo_id = criar_processo({"nome": "Processo de teste"}, caminho_banco=caminho_db)

    salvar_preco_item(processo_id, 1, quantidade=10, preco_unitario=25.5, caminho_banco=caminho_db)
    salvar_preco_item(processo_id, 1, quantidade=12, preco_unitario=30.0, caminho_banco=caminho_db)

    precos = obter_precos_item(processo_id, caminho_banco=caminho_db)
    assert len(precos) == 1  # não duplicou linha
    assert precos[1]["quantidade"] == 12
    assert precos[1]["preco_unitario"] == 30.0


def test_salvar_preco_item_aceita_so_um_campo_de_cada_vez(caminho_db):
    # Padrão de preenchimento real: a pessoa digita quantidade primeiro,
    # sai do campo (salva), depois digita preço (salva de novo) -- cada
    # PATCH manda só o que está preenchido no momento, o outro fica None.
    processo_id = criar_processo({"nome": "Processo de teste"}, caminho_banco=caminho_db)

    salvar_preco_item(processo_id, 1, quantidade=10, preco_unitario=None, caminho_banco=caminho_db)
    linha = salvar_preco_item(processo_id, 1, quantidade=10, preco_unitario=25.5, caminho_banco=caminho_db)

    assert linha["quantidade"] == 10
    assert linha["preco_unitario"] == 25.5


def test_obter_precos_item_indexado_por_numero_item(caminho_db):
    processo_id = criar_processo({"nome": "Processo de teste"}, caminho_banco=caminho_db)
    salvar_preco_item(processo_id, 1, quantidade=1, preco_unitario=1.0, caminho_banco=caminho_db)
    salvar_preco_item(processo_id, 5, quantidade=5, preco_unitario=5.0, caminho_banco=caminho_db)

    precos = obter_precos_item(processo_id, caminho_banco=caminho_db)
    assert set(precos.keys()) == {1, 5}


def test_obter_precos_item_processo_sem_preco_devolve_dict_vazio(caminho_db):
    processo_id = criar_processo({"nome": "Processo de teste"}, caminho_banco=caminho_db)
    assert obter_precos_item(processo_id, caminho_banco=caminho_db) == {}


# ---------- Repositório: marca/fabricante/modelo (Fase 4, Camada 1 da
# minuta, 19/08/2026 -- mesmo ponto de entrada por item de preco_item) ----------


def test_salvar_preco_item_com_marca_fabricante_modelo(caminho_db):
    processo_id = criar_processo({"nome": "Processo de teste"}, caminho_banco=caminho_db)

    linha = salvar_preco_item(
        processo_id, 1, quantidade=5, preco_unitario=200.0,
        marca="MarcaX", fabricante="FabricanteY", modelo="ModeloZ",
        caminho_banco=caminho_db,
    )

    assert linha["marca"] == "MarcaX"
    assert linha["fabricante"] == "FabricanteY"
    assert linha["modelo"] == "ModeloZ"


def test_salvar_preco_item_marca_fabricante_modelo_sao_opcionais(caminho_db):
    # Nem todo item exige marca/fabricante/modelo (ex.: serviço, não
    # produto) -- os três ficam None sem quebrar nada.
    processo_id = criar_processo({"nome": "Processo de teste"}, caminho_banco=caminho_db)

    linha = salvar_preco_item(processo_id, 1, quantidade=1, preco_unitario=10.0, caminho_banco=caminho_db)

    assert linha["marca"] is None
    assert linha["fabricante"] is None
    assert linha["modelo"] is None


def test_salvar_preco_item_atualiza_marca_sem_perder_quantidade_ja_salva(caminho_db):
    # Padrão real de preenchimento: cada blur reenvia o estado ATUAL da
    # linha inteira (não só o campo editado) -- o JS que garante isso;
    # aqui confere que o repositório aplica exatamente o que recebe.
    processo_id = criar_processo({"nome": "Processo de teste"}, caminho_banco=caminho_db)

    salvar_preco_item(processo_id, 1, quantidade=5, preco_unitario=200.0, caminho_banco=caminho_db)
    linha = salvar_preco_item(
        processo_id, 1, quantidade=5, preco_unitario=200.0, marca="MarcaX", caminho_banco=caminho_db
    )

    assert linha["quantidade"] == 5
    assert linha["preco_unitario"] == 200.0
    assert linha["marca"] == "MarcaX"


# ---------- Repositório: validade da proposta (processo.validade_proposta,
# Fase 4, Camada 1 da minuta, 19/08/2026) ----------


def test_atualizar_validade_proposta_salva_texto_livre(caminho_db):
    processo_id = criar_processo({"nome": "Processo de teste"}, caminho_banco=caminho_db)

    atualizar_validade_proposta(processo_id, "60 dias", caminho_banco=caminho_db)

    processo = obter_processo(processo_id, caminho_banco=caminho_db)
    assert processo is not None
    assert processo["validade_proposta"] == "60 dias"


def test_atualizar_validade_proposta_aceita_none_pra_limpar_o_campo(caminho_db):
    processo_id = criar_processo({"nome": "Processo de teste"}, caminho_banco=caminho_db)
    atualizar_validade_proposta(processo_id, "60 dias", caminho_banco=caminho_db)

    atualizar_validade_proposta(processo_id, None, caminho_banco=caminho_db)

    processo = obter_processo(processo_id, caminho_banco=caminho_db)
    assert processo is not None
    assert processo["validade_proposta"] is None


def test_atualizar_validade_proposta_processo_inexistente_levanta_erro(caminho_db):
    with pytest.raises(RegistroNaoEncontradoError):
        atualizar_validade_proposta(999999, "60 dias", caminho_banco=caminho_db)


# ---------- Geração do XLSX (função pura) ----------


def _catalogo(*numeros_e_textos: tuple[int, str]) -> list[dict]:
    return [{"numero": n, "texto_bruto": t, "pagina": 1, "localizador": "página 1"} for n, t in numeros_e_textos]


def test_gerar_planilha_sem_catalogo_levanta_erro():
    with pytest.raises(SemCatalogoError, match="1"):
        gerar_planilha_preco({"id": 1, "nome": "Processo X"}, [], {})


def test_gerar_planilha_completa_calcula_total_por_item_e_geral():
    catalogo = _catalogo((1, "Cadeira de rodas"), (2, "Muleta axilar"))
    precos = {
        1: {"quantidade": 5, "preco_unitario": 100.0},
        2: {"quantidade": 20, "preco_unitario": 15.0},
    }

    pasta = gerar_planilha_preco({"id": 1, "nome": "Processo X"}, catalogo, precos)
    planilha = pasta.active
    assert planilha is not None

    assert planilha["A2"].value == 1
    assert planilha["B2"].value == "Cadeira de rodas"
    assert planilha["C2"].value == 5
    assert planilha["D2"].value == 100.0
    assert planilha["E2"].value == 500.0  # 5 * 100

    assert planilha["E3"].value == 300.0  # 20 * 15

    # Linha de total geral, logo depois do último item.
    assert planilha["D4"].value == "Total geral:"
    assert planilha["E4"].value == 800.0  # 500 + 300


def test_gerar_planilha_item_sem_preco_fica_vazio_sem_travar_os_demais():
    catalogo = _catalogo((1, "Item sem preço"), (2, "Item com preço"))
    precos = {2: {"quantidade": 3, "preco_unitario": 10.0}}

    pasta = gerar_planilha_preco({"id": 1, "nome": "Processo X"}, catalogo, precos)
    planilha = pasta.active
    assert planilha is not None

    # Item 1: sem preco_item nenhum salvo -- tudo vazio, nada inventado.
    assert planilha["C2"].value is None
    assert planilha["D2"].value is None
    assert planilha["E2"].value is None

    # Item 2 continua calculado normalmente -- item sem dado não trava os outros.
    assert planilha["C3"].value == 3
    assert planilha["D3"].value == 10.0
    assert planilha["E3"].value == 30.0


def test_gerar_planilha_item_com_so_quantidade_nao_calcula_total_parcial():
    # Meio dado não vira meia conta -- sem os dois números, não dá pra
    # inventar o preço total.
    catalogo = _catalogo((1, "Item parcial"))
    precos = {1: {"quantidade": 10, "preco_unitario": None}}

    pasta = gerar_planilha_preco({"id": 1, "nome": "Processo X"}, catalogo, precos)
    planilha = pasta.active
    assert planilha is not None

    assert planilha["C2"].value == 10
    assert planilha["D2"].value is None
    assert planilha["E2"].value is None


def test_gerar_planilha_devolve_xlsx_reabrivel():
    # Prova real de que é um XLSX válido, não só bytes com a extensão certa.
    import io

    catalogo = _catalogo((1, "Item único"))
    precos = {1: {"quantidade": 2, "preco_unitario": 50.0}}
    pasta = gerar_planilha_preco({"id": 1, "nome": "Processo X"}, catalogo, precos)

    buffer = io.BytesIO()
    pasta.save(buffer)
    buffer.seek(0)

    reaberta = load_workbook(buffer)
    planilha = reaberta.active
    assert planilha is not None
    assert planilha["A1"].value == "Item"
    assert planilha["A2"].value == 1


# ---------- Teste de ouro: extração do catálogo contra edital real ----------


@pytest.mark.skipif(not CAMINHO_PAULINIA.exists(), reason="PDF real de Paulínia não está disponível localmente")
def test_extracao_do_catalogo_contra_edital_real_de_paulinia():
    # 100% determinístico (extrair_texto + fatiar_por_item), zero IA --
    # roda offline, sem gastar cota do Gemini. Mesmo edital já usado como
    # golden test em outras partes do projeto (Paulínia, processo id 1/7
    # do banco real).
    documento = extrair_texto(str(CAMINHO_PAULINIA))
    resultado = localizar_documento_com_tabela([documento])

    assert resultado is not None, "edital de Paulínia deveria ter uma tabela de itens reconhecível"
    _, itens = resultado

    assert len(itens) > 0
    # Números sequenciais a partir de 1, sem furo -- mesma garantia que
    # fatiar_por_item já oferece (tests/test_tabela_itens.py).
    assert [item.numero for item in itens] == list(range(1, len(itens) + 1))
    # Cada item tem texto de verdade, não vazio -- e página válida.
    for item in itens:
        assert item.texto.strip() != ""
        assert item.pagina is not None and item.pagina >= 1


@pytest.mark.skipif(not CAMINHO_PAULINIA.exists(), reason="PDF real de Paulínia não está disponível localmente")
def test_geracao_da_planilha_contra_catalogo_real_de_paulinia_com_preco_de_teste():
    # Fecha o ciclo completo com dado real: extrai o catálogo de verdade,
    # preenche preço sintético (não é dado real de negócio, só validação),
    # gera a planilha e confere que reabre e calcula certo.
    documento = extrair_texto(str(CAMINHO_PAULINIA))
    _, itens = localizar_documento_com_tabela([documento])  # type: ignore[misc]

    catalogo = [
        {"numero": item.numero, "texto_bruto": item.texto, "pagina": item.pagina, "localizador": item.localizador}
        for item in itens
    ]
    # Preço de teste só no primeiro item -- confirma que o resto some
    # em branco/destacado, sem travar a geração.
    precos = {itens[0].numero: {"quantidade": 10, "preco_unitario": 99.9}}

    pasta = gerar_planilha_preco({"id": 999, "nome": "PM Paulínia (teste)"}, catalogo, precos)
    planilha = pasta.active
    assert planilha is not None

    assert planilha.max_row == len(catalogo) + 2  # cabeçalho + itens + total geral
    assert planilha["C2"].value == 10
    assert planilha["D2"].value == 99.9
    assert planilha["E2"].value == 999.0
